import os
import time
import requests
from bs4 import BeautifulSoup
from collections import defaultdict
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


class ScrapePowerOutageLive:

    def __init__(self, country, country_url):
        self.base_url = "https://poweroutage.live"
        self.country_url = country_url
        self.url = self.base_url + self.country_url
        self.country = country


    def fetch(self):
        response = requests.get(self.url)
        if response.status_code != 200:
            print(f"Failed to access {self.url}")
            exit()
        else:
            return response

    def get_links(self, response):
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            links = [a['href'] for a in soup.select('.links-list a')]
            return links
        else:
            print(f"Failed to fetch the page. Status Code: {response.status_code}")
            exit()

    def scrape(self):
        response = self.fetch()
        links = self.get_links(response)
        for link in links:
            url = self.base_url + link
            data = self.process(url)
            times, state = self.find_outage_times(data)
            if len(state) > 31:
                state = state[:31]
            self.generate_xls(state, self.country, times)


    def find_outage_times(self, data):
        times = dict()
        date, state = data
        for day in date.keys():
            times[day] = []
            hours, outages = date[day]
            for i, outage in enumerate(outages):
                # if outage == "●":
                if outage == "✕":
                    times[day].append(hours[i])
        return times, state

    def process(self, url):
        response = requests.get(url)
        state = url.split("/")[-1][:-1]
        data = defaultdict(tuple)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            schedule_for_today = soup.find('h2', string="Schedule for today")
            if schedule_for_today:
                table_today = schedule_for_today.find_next('table')
                if table_today:
                    rows = table_today.find_all('tr')
                    today_time = rows[0]
                    if today_time:
                        today_times = [td.get_text(strip=True) for td in today_time.find_all('td')]
                    else:
                        print("No <tr> found in the table.")
                    today_outage = rows[1]
                    if today_outage:
                        today_outage_marks = [td.get_text(strip=True) for td in today_outage.find_all('td')]
                    else:
                        print("No <tr> found in the table.")
                    data[today_times[0]] = (today_times[1:], today_outage_marks[1:])
                else:
                    print("No table found after the H2 tag.")
            else:
                print("No H2 tag with 'Schedule for today' found.")
            schedule_for_tomorrow = soup.find('h2', string="Schedule for tomorrow")
            if schedule_for_tomorrow:
                table_tomorrow = schedule_for_tomorrow.find_next('table')
                if table_tomorrow:
                    rows = table_tomorrow.find_all('tr')
                    tomorrow_time = rows[0]
                    if tomorrow_time:
                        tomorrow_times = [td.get_text(strip=True) for td in tomorrow_time.find_all('td')]
                    else:
                        print("No <tr> found in the table.")
                    tomorrow_outage = rows[1]
                    if tomorrow_outage:
                        tomorrow_outage_marks = [td.get_text(strip=True) for td in tomorrow_outage.find_all('td')]
                    else:
                        print("No <tr> found in the table.")
                    data[tomorrow_times[0]] = (tomorrow_times[1:], tomorrow_outage_marks[1:])
                else:
                    print("No table found after the H2 tag.")
            else:
                print("No H2 tag with 'Schedule for tomorrow' found.")
            return data, state
        else:
            print(f"Failed to fetch the page. Status Code: {response.status_code}")

    def generate_xls(self, state, country, times):

        file_name = country + "_" + datetime.today().strftime("%Y-%m-%d") + ".xlsx"

        if os.path.exists(file_name):
            book = load_workbook(file_name)
            if state not in book.sheetnames:
                book.create_sheet(state)
            book.save(file_name)
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Write the DataFrame to the specified sheet
                dates = list(times.keys())
                hours = [f"{i:02d}" for i in range(24)]
                df = pd.DataFrame(index=hours, columns=dates)
                for date, hours in times.items():
                    for hour in hours:
                        df.at[hour, date] = "X"
                df.to_excel(writer, sheet_name=state, index=True)
        else:
            dates = list(times.keys())
            hours = [f"{i:02d}" for i in range(24)]
            df = pd.DataFrame(index=hours, columns=dates)
            for date, hours in times.items():
                for hour in hours:
                    df.at[hour, date] = "X"
            df.to_excel(file_name, sheet_name=state, index=True)

def get_countries():
    url = "https://poweroutage.live/"
    response = requests.get(url)
    res = []
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        links = [a['href'] for a in soup.select('.links-list li')]
        for link in links:
            res.append("/" + link)
    return res

countries = get_countries()

country_urls = ["/pk"]
for country_url in country_urls:
    country = country_url[1:]
    scrapePowerOutageLive = ScrapePowerOutageLive(country, country_url)
    scrapePowerOutageLive.scrape()
    time.sleep(1)





